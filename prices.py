from scrapy import Spider
from scrapy.http import Request
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from time import sleep
from urllib.parse import urljoin
from openpyxl import load_workbook
from openpyxl.utils import coordinate_from_string


class PricesSpider(Spider):
    name = 'prices'
    allowed_domains = ["web_to_scrape.com"]

    def start_requests(self):

        # Opens Excel workbook then particular worksheet with parts numbers column
        dest_filename = 'test.xlsx'
        wb = load_workbook(dest_filename)
        ws = wb['test']
        cell_range = ws['A1':'A4']

        # Loads Chrome driver with login page as first page and then logs in
        scrapy_url = 'web_to_scrape.com/login'
        self.driver = webdriver.Chrome()
        self.driver.get(scrapy_url)
        sleep(2)

        self.driver.find_element_by_id("login").send_keys('Foo')
        self.driver.find_element_by_id("password").send_keys('Foo?')
        self.driver.find_element_by_id("login_button").click()
        sleep(5)

        # Iterates through cells with parts
        for part in cell_range:

            # Gets row index form part's cell to put result in same row
            a = (str(part[0])[-5:-1].replace(".", '').replace("'", ''))
            row = (coordinate_from_string(str(a))[1])

            # Goes to part webpage
            part_url = urljoin('http://web_to_scrape.com/', part[0].value)
            self.driver.get(part_url)

            # Checks if part's cell is empty or no, if empty goes to next one
            if part[0].value is not None:

                # Gets the part prices without currency code and part no to check if the same
                try:
                    price = self.driver.find_element_by_id("customer_price").text.replace('EUR', ' ').replace('.', '')
                    list_price = self.driver.find_element_by_id("list_price").text.replace('EUR', ' ').replace('.', '')
                    part_no = self.driver.find_element_by_class_name("product_id").text

                    # Writes to the excel file
                    ws.cell(row=row, column=3, value=price)
                    ws.cell(row=row, column=4, value=list_price)
                    ws.cell(row=row, column=5, value=part_no)
                    wb.save(filename=dest_filename)

                # If part no is wrong, puts error massage
                except NoSuchElementException:
                        ws.cell(row=row, column=3, value='Wrong part no')
                        wb.save(filename=dest_filename)

                yield Request(part_url, callback=self.parse_price)

            else:
                pass

    def parse_price(self, response):
        pass
